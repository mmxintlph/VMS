import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = openpyxl.Workbook()

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── SHEET 1: CheckIns ──────────────────────────────────────────
ws1 = wb.active
ws1.title = "CheckIns"

HEADERS = ["Timestamp","Date","Shift","Employee Name","Email",
           "Department","Mood","Mood Score","Note","Week Number","Month"]

hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1E3A5F")
efill = PatternFill("solid", fgColor="EBF3FB")
blue  = Font(name="Arial", size=10, color="0070C0")
black = Font(name="Arial", size=10, color="000000")

for col, h in enumerate(HEADERS, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = hf; c.fill = hfill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws1.row_dimensions[1].height = 30

example_data = [
    datetime.datetime(2025, 7, 17, 8, 32, 0),
    "=INT(A2)",
    "Start of Shift",
    "Juan Cruz",
    "jcruz@company.com",
    "Customer Service",
    "Focused",
    4,
    "Ready for the day.",
    '=IFERROR(WEEKNUM(B2,2),"")',
    '=IFERROR(TEXT(B2,"MMMM YYYY"),"")',
]

for col, val in enumerate(example_data, 1):
    c = ws1.cell(row=2, column=col, value=val)
    c.fill = efill; c.border = border
    c.alignment = Alignment(vertical="center")
    c.font = black if (isinstance(val, str) and val.startswith("=")) else blue

ws1.cell(row=2, column=1).number_format = "YYYY-MM-DD HH:MM:SS"
ws1.cell(row=2, column=2).number_format = "YYYY-MM-DD"

col_widths = [22,14,16,22,30,20,14,12,40,13,16]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"

muted = Font(name="Arial", size=9, color="7F7F7F", italic=True)
ws1["A4"] = "LEGEND"
ws1["A4"].font = Font(name="Arial", bold=True, size=9, color="1E3A5F")
notes = [
    (5, "Blue text = values written by the MoodMate app (do not edit manually)"),
    (6, "Black text = auto-calculated formulas"),
    (7, "Row 2 is an example entry — the app appends real data in the next available row."),
]
for row, txt in notes:
    ws1.cell(row=row, column=1, value=txt).font = muted
    ws1.merge_cells(f"A{row}:K{row}")

# ── SHEET 2: Dashboard ─────────────────────────────────────────
ws2 = wb.create_sheet("Dashboard")

ws2["A1"] = "MoodMate — Summary Dashboard"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1E3A5F")
ws2.merge_cells("A1:F1")
ws2["A2"] = "Auto-calculated from the CheckIns tab. Add data there; this sheet updates automatically."
ws2["A2"].font = Font(name="Arial", size=9, italic=True, color="7F7F7F")
ws2.merge_cells("A2:F2")

sf = Font(name="Arial", bold=True, size=10, color="FFFFFF")
sfill = PatternFill("solid", fgColor="2E75B6")
vf = Font(name="Arial", bold=True, size=14, color="1E3A5F")
lf = Font(name="Arial", size=9, color="595959")

# KPI section header
c = ws2.cell(row=4, column=1, value="OVERALL KPIs")
c.font = sf; c.fill = sfill
c.alignment = Alignment(horizontal="center")
ws2.merge_cells("A4:F4")
ws2.row_dimensions[4].height = 20

kpis = [
    (1, "Total Check-ins",     '=IFERROR(COUNTA(CheckIns!A2:A10000)-3,0)',                               "0"),
    (2, "Avg Mood Score",      '=IFERROR(AVERAGEIF(CheckIns!G2:G10000,"<>",CheckIns!H2:H10000),0)',      "0.00"),
    (3, "% Positive (≥4)",     '=IFERROR(COUNTIF(CheckIns!H2:H10000,">=4")/COUNTA(CheckIns!H2:H10000),0)', "0%"),
    (4, "% At Risk (≤2)",      '=IFERROR(COUNTIF(CheckIns!H2:H10000,"<=2")/COUNTA(CheckIns!H2:H10000),0)', "0%"),
    (5, "Unique Employees",    '=IFERROR(SUMPRODUCT(1/COUNTIF(CheckIns!D2:D100,CheckIns!D2:D100)),0)',   "0"),
    (6, "Last Entry Date",     '=IFERROR(TEXT(MAX(CheckIns!B2:B10000),"YYYY-MM-DD"),"")',                "@"),
]

for col, label, formula, fmt in kpis:
    lc = ws2.cell(row=5, column=col, value=label)
    lc.font = lf; lc.alignment = Alignment(wrap_text=True, horizontal="center")
    vc = ws2.cell(row=6, column=col, value=formula)
    vc.font = vf; vc.number_format = fmt
    vc.alignment = Alignment(horizontal="center")
    vc.fill = PatternFill("solid", fgColor="EBF3FB")
    vc.border = Border(bottom=Side(style="medium", color="2E75B6"))

ws2.row_dimensions[5].height = 30
ws2.row_dimensions[6].height = 28

# Mood breakdown — headers first (A8:C8), then section merge label on A7
ws2["A7"] = "MOOD BREAKDOWN"
ws2["A7"].font = sf; ws2["A7"].fill = sfill
ws2["A7"].alignment = Alignment(horizontal="center")
ws2.merge_cells("A7:C7")
ws2.row_dimensions[7].height = 20

for col, h in enumerate(["Mood","Count","% of Total"], 1):
    c = ws2.cell(row=8, column=col, value=h)
    c.font = Font(name="Arial", bold=True, size=9, color="1E3A5F")
    c.fill = PatternFill("solid", fgColor="D6E4F0")
    c.alignment = Alignment(horizontal="center")
    c.border = border

moods = [
    ("Energized", "22D3EE"),
    ("Focused",   "6EE7B7"),
    ("Neutral",   "A78BFA"),
    ("Stressed",  "FCD34D"),
    ("Burned Out","F87171"),
]
for i, (mood, color) in enumerate(moods):
    row = 9 + i
    mc = ws2.cell(row=row, column=1, value=mood)
    mc.font = Font(name="Arial", size=10, bold=True)
    mc.fill = PatternFill("solid", fgColor=color + "44")
    mc.border = border
    cnt = ws2.cell(row=row, column=2, value=f'=IFERROR(COUNTIF(CheckIns!G2:G10000,"{mood}"),0)')
    cnt.font = Font(name="Arial", size=10); cnt.border = border; cnt.alignment = Alignment(horizontal="center")
    pct = ws2.cell(row=row, column=3, value=f'=IFERROR(B{row}/COUNTA(CheckIns!G2:G10000),0)')
    pct.font = Font(name="Arial", size=10); pct.number_format = "0.0%"
    pct.border = border; pct.alignment = Alignment(horizontal="center")

# Score reference (cols E-F)
ws2["E7"] = "SCORE REFERENCE"
ws2["E7"].font = sf; ws2["E7"].fill = sfill
ws2["E7"].alignment = Alignment(horizontal="center")
ws2.merge_cells("E7:F7")

for col, h in enumerate(["Score","Mood"], 1):
    c = ws2.cell(row=8, column=4+col, value=h)
    c.font = Font(name="Arial", bold=True, size=9, color="1E3A5F")
    c.fill = PatternFill("solid", fgColor="D6E4F0")
    c.alignment = Alignment(horizontal="center"); c.border = border

ref = [("5","Energized ⚡"),("4","Focused 🎯"),("3","Neutral 😐"),("2","Stressed 😰"),("1","Burned Out 😔")]
for i, (sc, lbl) in enumerate(ref):
    ws2.cell(row=9+i, column=5, value=sc).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=9+i, column=6, value=lbl).font = Font(name="Arial", size=10)

col_letters = ["A","B","C","D","E","F"]
    for col, w in zip(col_letters, [22,12,14,4,10,20]):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = "A4"

# ── SHEET 3: Config ────────────────────────────────────────────
ws3 = wb.create_sheet("Config")
ws3["A1"] = "MoodMate Configuration — IT Admin"
ws3["A1"].font = Font(name="Arial", bold=True, size=13, color="1E3A5F")
ws3.merge_cells("A1:C1")
ws3["A2"] = "Fill in the yellow cells below, then update moodmate.html with the same values."
ws3["A2"].font = Font(name="Arial", size=9, italic=True, color="7F7F7F")
ws3.merge_cells("A2:C2")

cfg_rows = [
    ("Setting", "Value", "Where to find it"),
    ("Azure AD Client ID",       "REPLACE_WITH_CLIENT_ID",                          "Azure Portal → App Registrations → your app → Overview"),
    ("Azure AD Tenant ID",       "REPLACE_WITH_TENANT_ID",                          "Azure Portal → Azure Active Directory → Overview"),
    ("SharePoint Site URL",      "https://yourcompany.sharepoint.com/sites/HR",     "Copy from your browser when on the SharePoint site"),
    ("SharePoint Drive Name",    "Documents",                                        "Usually 'Documents' — check SharePoint site contents"),
    ("Excel File Path in Drive", "/MoodMate/MoodMate_Database.xlsx",                "Path inside the Documents library where this file is uploaded"),
    ("CheckIns Sheet Name",      "CheckIns",                                        "Must match the tab name exactly (case-sensitive)"),
]

yfill = PatternFill("solid", fgColor="FFF2CC")
for r, row in enumerate(cfg_rows, 3):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        if r == 3:
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
        else:
            cell.font = Font(name="Arial", size=10, color="0070C0" if c == 2 else "000000")
            if c == 2: cell.fill = yfill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws3.row_dimensions[r].height = 22

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 50
ws3.column_dimensions["C"].width = 44

out = "/sessions/peaceful-gifted-ride/mnt/outputs/MoodMate_Database.xlsx"
wb.save(out)
print("Saved:", out)
